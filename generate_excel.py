from openpyxl.compat import range
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl import *
import datetime

# general config
TEMPLATE_EXCEL = 'template_excel.xlsx'
WRITE_SHEET_NAME = 'Pending Measurements'

# read template related
MAX_STATION_NUMBER = 100
STATION_DATA_ROW_START_NUMBER = 2
TEMPLATE_STATION_ID_COLUMN_NUMBER = 2
TEMPLATE_STATION_START_VAL_COLUMN_NUMBER = 3
TEMPLATE_STATION_END_VAL_COLUMN_NUMBER = 4

# write template related
STATION_TITLE_TEMPLATE = 'ST{0}'
M_RAIL_TITLE = 'M Rail'
S_RAIL_TITLE = 'S Rail'
M_RAIL_ROW_TEMPLATE = 'ST{0}_M_{1:02d}'
S_RAIL_ROW_TEMPLATE = 'ST{0}_S_{1:02d}'

def parse_expected_stations():
    template_wb = load_workbook(TEMPLATE_EXCEL)
    input_ws = template_wb["UI"]
    stations = []

    ids = [input_ws.cell(column=TEMPLATE_STATION_ID_COLUMN_NUMBER, row=i).value 
           for i in range(STATION_DATA_ROW_START_NUMBER,MAX_STATION_NUMBER) 
           if input_ws.cell(column=TEMPLATE_STATION_END_VAL_COLUMN_NUMBER, row=i).value is not None]

    start_numbers = [input_ws.cell(column=TEMPLATE_STATION_START_VAL_COLUMN_NUMBER, row=i).value 
                     for i in range(STATION_DATA_ROW_START_NUMBER,MAX_STATION_NUMBER) 
                     if input_ws.cell(column=TEMPLATE_STATION_END_VAL_COLUMN_NUMBER, row=i).value is not None]

    end_numbers = [input_ws.cell(column=TEMPLATE_STATION_END_VAL_COLUMN_NUMBER, row=i).value 
                   for i in range(STATION_DATA_ROW_START_NUMBER,MAX_STATION_NUMBER) 
                   if input_ws.cell(column=TEMPLATE_STATION_END_VAL_COLUMN_NUMBER, row=i).value is not None]

    if not (len(ids) == len(start_numbers) == len(end_numbers)):
        print(
            """
            Input error. I saw {0} ids, {1} start values, and {2} end values. 
            I need an an equal number of these 3. Check the input sheet.
            """.format(len(ids), len(start_numbers), len(end_numbers)))
        exit(1)

    # we have checked equality, so we're good to slam these arrays together to get tuples
    for _id, start, end in zip(ids, start_numbers, end_numbers):
        # just in case someone derped and reversed which was which for start and end
        if start > end:
            stations.append((_id, end, start))
        else:
            stations.append((_id, start, end))
        
    return stations

# set cell
def s_c(ws, x, y, val, border = None, shading = None):
    if border is not None:
        ws.cell(row=y, column=x).border = border

    if shading is not None:
        ws.cell(row=y, column=x).fill = shading

    # hack, don't feel like adding a parameter for style
    if y == 1 and val:
        ws.cell(row=y, column=x).style = 'Headline 1'

    ws.cell(row=y, column=x).value = val

def write_expected_stations(stations):
    dest_filename = 'generated_{:%m%d%Y%H%M}.xlsx'.format(datetime.datetime.now())
    wb = Workbook()

    ws = wb.active
    ws.title = WRITE_SHEET_NAME

    left_thick = Border(left=Side(style='medium'))
    right_thick = Border(right=Side(style='medium'))

    even_header_fill = PatternFill(patternType='solid', fgColor=Color(rgb='66A5AD'))
    odd_header_fill = PatternFill(patternType='solid', fgColor=Color(rgb='BA5536'))

    # easy access for h_clr[N % 2]
    h_clr = [
        even_header_fill,
        odd_header_fill
    ]

    # n = station id
    for N, station in enumerate(stations):
        # station tuple definition ref
        # 0 = id
        # 1 = start value
        # 2 = end value
        station_id = station[0]
        start = station[1]
        end = station[2]

        # all formulas are written in x, y format, including cell() calls
        # insert station header
        # 0N + 1, row = 1
        s_c(ws, 9 * N + 1, 1, STATION_TITLE_TEMPLATE.format(N), left_thick)
        s_c(ws, 9 * N + 8, 1, '', right_thick)

        # insert column headers, row = 2
        s_c(ws, 9 * N + 1, 2, M_RAIL_TITLE, left_thick, h_clr[N % 2])
        s_c(ws, 9 * N + 2, 2, 'X', None, h_clr[N % 2])
        s_c(ws, 9 * N + 3, 2, 'Y', None, h_clr[N % 2])
        s_c(ws, 9 * N + 4, 2, 'Z', None, h_clr[N % 2])
        s_c(ws, 9 * N + 5, 2, S_RAIL_TITLE, None, h_clr[N % 2])
        s_c(ws, 9 * N + 6, 2, 'X', None, h_clr[N % 2])
        s_c(ws, 9 * N + 7, 2, 'Y', None, h_clr[N % 2])
        s_c(ws, 9 * N + 8, 2, 'Z', right_thick, h_clr[N % 2])

        # row = 3 + index
        # insert station row templates between start and end indexes
        for index, m in enumerate(range(start, end)):
            y = index + 3
            s_c(ws, 9 * N + 1, y, M_RAIL_ROW_TEMPLATE.format(N, m), left_thick)
            s_c(ws, 9 * N + 5, y, S_RAIL_ROW_TEMPLATE.format(N, m))

            s_c(ws, 9 * N + 8, y, '', right_thick)

        wb.save(filename = dest_filename)

if __name__ == '__main__':
    expected_stations = parse_expected_stations()
    write_expected_stations(expected_stations)  

    exit(0)